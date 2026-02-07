#!/usr/bin/env python3
"""Extract text from common file formats."""

import csv
import json
from pathlib import Path
from typing import Optional

# Base formats (always available)
SUPPORTED_EXT = {".txt", ".json", ".csv", ".md", ".rtf", ".xml"}


def _read_text(path: Path, encoding: str = "utf-8") -> str:
    """Read text with encoding fallback."""
    try:
        return path.read_text(encoding=encoding, errors="replace").strip()
    except Exception:
        return path.read_text(encoding="latin-1", errors="replace").strip()


def _load_docx(path: Path) -> Optional[str]:
    """Word .docx – requires: pip install python-docx"""
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return None
    except Exception:
        return None


def _load_xlsx(path: Path) -> Optional[str]:
    """Excel .xlsx – requires: pip install openpyxl"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    lines.append(" | ".join(cells))
        wb.close()
        return "\n".join(lines) if lines else None
    except ImportError:
        return None
    except Exception:
        return None


def _load_xls(path: Path) -> Optional[str]:
    """Excel .xls (legacy) – requires: pip install xlrd"""
    try:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        lines: list[str] = []
        for sheet in wb.sheets():
            for row_idx in range(sheet.nrows):
                cells = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                if any(c.strip() for c in cells):
                    lines.append(" | ".join(cells))
        return "\n".join(lines) if lines else None
    except ImportError:
        return None
    except Exception:
        return None


def _load_pdf(path: Path) -> Optional[str]:
    """PDF – requires: pip install pypdf"""
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        text = "\n".join(p.extract_text() or "" for p in reader.pages).strip()
        if text:
            return text
        return _ocr_pdf(path)
    except ImportError:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(path)
            text = "\n".join(p.extract_text() or "" for p in reader.pages).strip()
            if text:
                return text
            return _ocr_pdf(path)
        except ImportError:
            return None
    except Exception:
        return None


def _ocr_pdf(path: Path) -> Optional[str]:
    """OCR for scanned PDFs – requires: pytesseract + pdf2image + poppler."""
    try:
        import pytesseract
        from pdf2image import convert_from_path
    except ImportError:
        return None
    try:
        images = convert_from_path(str(path))
        pages = [pytesseract.image_to_string(img).strip() for img in images]
        text = "\n".join(p for p in pages if p)
        return text if text else None
    except Exception:
        return None


def _load_image(path: Path) -> Optional[str]:
    """Image OCR – requires: pytesseract + pillow + Tesseract binary."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return None
    try:
        image = Image.open(path)
        return pytesseract.image_to_string(image).strip()
    except Exception:
        return None


def _load_odt(path: Path) -> Optional[str]:
    """OpenDocument Text .odt – requires: pip install odfpy"""
    try:
        from odf import text, teletype
        from odf.opendocument import load
        doc = load(str(path))
        paras = doc.getElementsByType(text.P)
        return "\n".join(teletype.extractText(p) for p in paras if teletype.extractText(p).strip())
    except ImportError:
        return None
    except Exception:
        return None


def _load_csv(path: Path) -> str:
    """CSV – standard library."""
    lines: list[str] = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.reader(f):
            if any(c.strip() for c in row):
                lines.append(" | ".join(row))
    return "\n".join(lines)


def extract_text(path: Path) -> Optional[str]:
    """
    Extract text from a file.
    Supports: .txt, .json, .csv, .md, .rtf, .xml, .docx, .xlsx, .xls, .pdf, .odt, images
    Returns None on unsupported format or failure.
    """
    ext = path.suffix.lower()
    if not path.is_file():
        return None

    # Text formats
    if ext in {".txt", ".md", ".rtf", ".xml"}:
        return _read_text(path)

    if ext == ".json":
        content = _read_text(path)
        # Pass JSON as-is (may contain IDP data)
        return content

    if ext == ".csv":
        return _load_csv(path)

    # Binary formats (optional)
    if ext == ".docx":
        return _load_docx(path)
    if ext == ".xlsx":
        return _load_xlsx(path)
    if ext == ".xls":
        return _load_xls(path)
    if ext == ".pdf":
        return _load_pdf(path)
    if ext == ".odt":
        return _load_odt(path)
    if ext in {".png", ".jpg", ".jpeg", ".tiff", ".bmp"}:
        return _load_image(path)

    return None


def get_supported_extensions() -> set[str]:
    """Return all supported extensions (including optional)."""
    ext = SUPPORTED_EXT.copy()
    try:
        from docx import Document
        ext.add(".docx")
    except ImportError:
        pass
    try:
        from openpyxl import load_workbook
        ext.add(".xlsx")
    except ImportError:
        pass
    try:
        import xlrd
        ext.add(".xls")
    except ImportError:
        pass
    try:
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader
        ext.add(".pdf")
    except ImportError:
        pass
    try:
        from odf.opendocument import load
        ext.add(".odt")
    except ImportError:
        pass
    try:
        import pytesseract
        from PIL import Image
        ext.update({".png", ".jpg", ".jpeg", ".tiff", ".bmp"})
    except ImportError:
        pass
    return ext
